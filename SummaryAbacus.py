import streamlit as st
import pandas as pd
from openpyxl import load_workbook


def run():
    st.title("Laporan RPL & ATM ABACUS")

    # --- Upload terpisah ---
    file_CIT_ABS = st.file_uploader("Upload laporan CIT (sheet CIT ABS)", type=["xlsx"])
    file_IC_ABS = st.file_uploader("Upload laporan IC (sheet IC ABS)", type=["xlsx"])
    file_wilayah_ABS = st.file_uploader(
        "Upload laporan Rencana & Wilayah(saldo Posisi Saldo ABS)", type=["xlsx"]
    )
    file_PosisiAwal_ABS = st.file_uploader(
        "Upload laporan Saldo awal ABS", type=["xlsx"]
    )

    if file_CIT_ABS and file_IC_ABS and file_wilayah_ABS and file_PosisiAwal_ABS:
        # --- 0. Detail CIT ---
        df_CIT = pd.read_excel(file_CIT_ABS, sheet_name="RECON", skiprows=11)
        # st.write(df_CIT.columns.tolist())
        # st.dataframe(df_CIT.head(30))
        cols_map = {
            "Unnamed: 1": "ATM ID",
            "Unnamed: 3": "Nama ATM",
            "Unnamed: 4": "Denom",
            "Unnamed: 6": "Tanggal Replenish",
            "Unnamed: 10": "JamRPL",
            "Unnamed: 5": "Cabang",
            "Unnamed: 7": "Total Kaset Sebelumnya",
            "Unnamed: 8": "Admin Sebelumnya",
            "CST 1": "Sisa uang Kaset 1",
            "CST 2": "Sisa uang Kaset 2",
            "CST 3": "Sisa uang Kaset 3",
            "CST 4": "Sisa uang Kaset 4",
            "RJT": "Sisa uang Kaset Reject",
            "CST 1.1": "Total uang keluar",
            "Unnamed: 26": "Total Kaset Terbaru",
            "Unnamed: 27": "Admin Terbaru",
        }

        available_cols = [c for c in cols_map.keys() if c in df_CIT.columns]
        df_selected_CIT = df_CIT[available_cols].copy()
        df_selected_CIT.rename(columns=cols_map, inplace=True)

        # Konversi numerik
        for col in df_selected_CIT.columns:
            df_selected_CIT[col] = pd.to_numeric(df_selected_CIT[col], errors="ignore")

        # Hitung turunan
        # Pastikan kolom kaset jadi numerik
        kaset_cols = [
            "Sisa uang Kaset 1",
            "Sisa uang Kaset 2",
            "Sisa uang Kaset 3",
            "Sisa uang Kaset 4",
            "Sisa uang Kaset Reject",
        ]

        for col in kaset_cols:
            if col in df_selected_CIT.columns:
                df_selected_CIT[col] = pd.to_numeric(
                    df_selected_CIT[col], errors="coerce"
                ).fillna(0)

        # Denom juga pastikan numerik
        df_selected_CIT["Denom"] = pd.to_numeric(
            df_selected_CIT["Denom"], errors="coerce"
        ).fillna(1)

        # Hitung Total Sisa Uang
        df_selected_CIT["Total Sisa Uang"] = (
            df_selected_CIT[kaset_cols].sum(axis=1) * df_selected_CIT["Denom"] * 1000
        )

        df_selected_CIT["Total Amount Uang Keluar"] = (
            df_selected_CIT.get("Total uang keluar", 0)
            * df_selected_CIT.get("Denom", 1)
            * 1000
        )

        # st.subheader("Detail ATM (CIT)")
        # st.dataframe(df_selected_CIT.head(10))

        # --- 1. Detail IC ---
        xls_IC = pd.ExcelFile(file_IC_ABS)
        sheet_names = xls_IC.sheet_names

        # Tampilkan pilihan ke user
        selected_sheet = st.selectbox(
            "Pilih tanggal dari laporan IC yang mau dibaca:", sheet_names
        )
        df_IC = pd.read_excel(file_IC_ABS, sheet_name=selected_sheet, skiprows=1)
        df_IC.rename(
            columns={
                "Unnamed: 0": "NO",
                "Unnamed: 1": "ID ATM",
                "Unnamed: 2": "Nama Lokasi",
            },
            inplace=True,
        )

        selected_cols = [
            "ID ATM",
            "Nama Lokasi",
            "Cabang",
            "Denom",
            "Limit",
            "Saldo",
            "Nominal",
        ]
        df_selected_IC = df_IC[[c for c in selected_cols if c in df_IC.columns]]

        for col in ["Denom", "Limit", "Saldo", "Nominal"]:
            if col in df_selected_IC.columns:
                df_selected_IC[col] = pd.to_numeric(
                    df_selected_IC[col], errors="coerce"
                )

        # st.subheader("Preview Data ATM (IC)")
        # st.dataframe(df_selected_IC.head(10))

        # --- 2. Rekap per Cabang (Rencana) ---
        df_cabang = pd.read_excel(file_wilayah_ABS, sheet_name="RENCANA", skiprows=1)
        cols_order = ["WSID", "LOKASI", "CABANG", "DENOM", "LIMIT", "KETERANGAN"]
        df_selected_cabang = df_cabang[
            [c for c in cols_order if c in df_cabang.columns]
        ]

        summary_cabang = (
            df_selected_cabang.groupby("CABANG")
            .agg(Jumlah_RPL=("WSID", "count"), Total_Limit=("LIMIT", "sum"))
            .reset_index()
        )

        # --- 3. Rekap per Wilayah ---
        wb = load_workbook(file_wilayah_ABS, data_only=True)
        ws = wb.active
        mapping = {
            "SUKABUMI": ["D4", "E4", "F4", "D5", "E5", "F5"],
            "SIDOARJO": ["G4", "H4", "I4", "G5", "H5", "I5"],
            "MALANG": ["J4", "K4", "L4", "J5", "K5", "L5"],
            "SURABAYA": ["M4", "N4", "O4", "M5", "N5", "O5"],
            "KEDIRI": ["P4", "Q4", "R4", "P5", "Q5", "R5"],
            "JAKARTA": ["D12", "E12", "F12", "D13", "E13", "F13"],
            "BANDUNG": ["G12", "H12", "I12", "G13", "H13", "I13"],
            "CIAMIS": ["J12", "K12", "L12", "J13", "K13", "L13"],
            "MEDAN": ["M12", "N12", "O12", "M13", "N13", "O13"],
            "BALI": ["P12", "Q12", "R12", "P13", "Q13", "R13"],
            "MAKASSAR": ["D20", "E20", "F20", "D21", "E21", "F21"],
            "BOGOR": ["G20", "H20", "I20", "G21", "H21", "I21"],
            "SUBANG": ["J20", "K20", "L20", "J21", "K21", "L21"],
        }
        data = []
        for wilayah, cells in mapping.items():
            values = [ws[c].value or 0 for c in cells]
            total_d100 = values[0] + values[3]
            total_d50 = values[1] + values[4]
            total_d20 = values[2] + values[5]
            awal_saldo = values[0] + values[1] + values[2]
            supply = values[3] + values[4] + values[5]
            total_saldo = total_d100 + total_d50 + total_d20
            row = {
                "Wilayah": wilayah,
                "Awal_saldo": awal_saldo,
                "Supply": supply,
                "Total Saldo": total_saldo,
            }
            data.append(row)
        df_wilayah = pd.DataFrame(data)

        # --- 3.1 Rekap posisi awal ---
        wb = load_workbook(file_PosisiAwal_ABS, data_only=True)
        ws = wb.active
        mapping = {
            "SUKABUMI": ["D4", "E4", "F4"],
            "SIDOARJO": ["G4", "H4", "I4"],
            "MALANG": ["J4", "K4", "L4"],
            "SURABAYA": ["M4", "N4", "O4"],
            "KEDIRI": ["P4", "Q4", "R4"],
            "JAKARTA": ["D12", "E12", "F12"],
            "BANDUNG": ["G12", "H12", "I12"],
            "CIAMIS": ["J12", "K12", "L12"],
            "MEDAN": ["M12", "N12", "O12"],
            "BALI": ["P12", "Q12", "R12"],
            "MAKASSAR": ["D20", "E20", "F20"],
            "BOGOR": ["G20", "H20", "I20"],
            "SUBANG": ["J20", "K20", "L20"],
        }
        data_realisasi = []
        for wilayah, cells in mapping.items():
            values = [ws[c].value or 0 for c in cells]
            total_d100 = values[0]
            total_d50 = values[1]
            total_d20 = values[2]
            posisi_saldo = values[0] + values[1] + values[2]
            row = {"Wilayah": wilayah, "Posisi Saldo": posisi_saldo}
            data_realisasi.append(row)
        df_realisasi = pd.DataFrame(data_realisasi)

        # --- 4. Gabungkan Wilayah dengan Rekap Cabang ---
        df_summary = df_wilayah.merge(
            summary_cabang, left_on="Wilayah", right_on="CABANG", how="left"
        )
        df_summary.rename(
            columns={
                "Jumlah_RPL": "Jumlah lokasi rencana pengisian",
                "Total_Limit": "Total nominal rencana pengisian",
            },
            inplace=True,
        )

        # --- 5. Tambahkan kolom realisasi dari IC ---
        summary_realisasi = (
            df_selected_IC.groupby("Cabang")
            .agg(
                Jumlah_Realisasi=("Nominal", lambda x: (x > 0).sum()),
                Nominal_Realisasi=("Nominal", "sum"),
            )
            .reset_index()
        )
        df_summary = df_summary.merge(
            summary_realisasi, left_on="Wilayah", right_on="Cabang", how="left"
        )
        df_summary.rename(
            columns={
                "Jumlah_Realisasi": "Jumlah lokasi realisasi pengisian",
                "Nominal_Realisasi": "Nominal realisasi pengisian",
            },
            inplace=True,
        )

        # --- 6. Tambahkan kolom Sisa uang dari CIT ---
        summary_sisa = (
            df_selected_CIT.groupby("Cabang")
            .agg(
                Jumlah_SisaUang=("Total Sisa Uang", lambda x: (x > 0).sum()),
                Nominal_SisaUang=("Total Sisa Uang", "sum"),
            )
            .reset_index()
        )
        df_summary = df_summary.merge(
            summary_sisa, left_on="Wilayah", right_on="Cabang", how="left"
        )
        df_summary.rename(
            columns={
                "Jumlah_SisaUang": "Jumlah lokasi sisa uang",
                "Nominal_SisaUang": "Nominal sisa uang",
            },
            inplace=True,
        )

        # --- 7. Hitung Saldo Akhir Rencana ---
        df_summary["Saldo Akhir Rencana"] = (
            df_summary["Total Saldo"].fillna(0)
            - df_summary["Total nominal rencana pengisian"].fillna(0)
            + df_summary["Nominal sisa uang"].fillna(0)
        )

        # --- 7.1 Hitung Saldo Akhir Realisasi ---
        df_summary["Saldo Akhir Realisasi"] = (
            df_summary["Total Saldo"].fillna(0)
            - df_summary["Nominal realisasi pengisian"].fillna(0)
            + df_summary["Nominal sisa uang"].fillna(0)
        )

        # st.write("Preview df_realisasi:", df_realisasi)
        # --- 7.2 Hitung Selisih nominal Rencana RPL ---
        df_summary = df_summary.merge(df_realisasi, on="Wilayah", how="left")
        df_summary["Selisih akhir Rencana"] = df_summary["Posisi Saldo"].fillna(
            0
        ) - df_summary["Saldo Akhir Rencana"].fillna(0)

        # --- 7.3 Hitung Selisih nominal realisasi RPL ---
        df_summary["Selisih akhir Realisasi"] = df_summary["Posisi Saldo"].fillna(
            0
        ) - df_summary["Saldo Akhir Realisasi"].fillna(0)

        # --- 7.4 Hitung Selisih Jumlah Replenishment ---
        df_summary["Selisih Jumlah Replenishment"] = df_summary[
            "Jumlah lokasi realisasi pengisian"
        ].fillna(0) - df_summary["Jumlah lokasi rencana pengisian"].fillna(0)

        # --- 7.5 Tambahkan kolom Remark per Wilayah ---
        # Pastikan WSID dan ID ATM jadi string tanpa .0
        wsid_per_cabang = (
            df_selected_cabang.groupby("CABANG")["WSID"]
            .apply(lambda x: set(x.dropna().astype(int).astype(str)))
            .to_dict()
        )
        idatm_per_cabang = (
            df_selected_IC.groupby("Cabang")["ID ATM"]
            .apply(lambda x: set(x.dropna().astype(int).astype(str)))
            .to_dict()
        )

        # Pastikan ID ATM di df_selected_IC jadi string bersih
        df_selected_IC["ID ATM"] = (
            df_selected_IC["ID ATM"]
            .astype(str)
            .str.strip()
            .str.replace(r"\.0$", "", regex=True)
        )

        remarks = []
        for idx, row in df_summary.iterrows():
            wilayah = row["Wilayah"]
            wsid_set = wsid_per_cabang.get(wilayah, set())
            idatm_set = idatm_per_cabang.get(wilayah, set())

            wsid_not_in_ic = wsid_set - idatm_set
            idatm_not_in_rencana = idatm_set - wsid_set

            if wsid_not_in_ic or idatm_not_in_rencana:
                remark_text = []
                if wsid_not_in_ic:
                    remark_text.append(
                        "WSID not in IC: " + ", ".join(sorted(wsid_not_in_ic))
                    )
                if idatm_not_in_rencana:
                    details = []
                    for atm_id in sorted(idatm_not_in_rencana):
                        row_ic = df_selected_IC.loc[df_selected_IC["ID ATM"] == atm_id]
                        if not row_ic.empty:
                            nominal_val = (
                                row_ic["Nominal"].iloc[0]
                                if pd.notna(row_ic["Nominal"].iloc[0])
                                else 0
                            )
                            details.append(f"{atm_id} (Nominal {int(nominal_val):,})")
                        else:
                            details.append(atm_id)
                    remark_text.append("ID ATM not in Rencana: " + ", ".join(details))
                remarks.append("; ".join(remark_text))
            else:
                remarks.append("Match")

        df_summary["Remark"] = remarks

        # --- 7.6 Hitung mismatch limit per terminal dari Detail IC ---
        df_selected_IC["Mismatch_Limit"] = df_selected_IC["Nominal"].fillna(
            0
        ) - df_selected_IC["Limit"].fillna(0)

        # Buat deskripsi mismatch per ATM (hindari NaN dengan pengecekan)
        def make_desc(row):
            if pd.notna(row["ID ATM"]) and row["Mismatch_Limit"] != 0:
                try:
                    atm_id = str(int(float(row["ID ATM"])))  # aman dari float NaN
                except Exception:
                    atm_id = str(row["ID ATM"])
                nominal_val = int(row["Nominal"]) if pd.notna(row["Nominal"]) else 0
                limit_val = int(row["Limit"]) if pd.notna(row["Limit"]) else 0
                return f"ATM {atm_id} mismatch: Realisasi {nominal_val:,} vs Limit {limit_val:,}"
            return None

        df_selected_IC["Mismatch_Desc"] = df_selected_IC.apply(make_desc, axis=1)

        # Gabungkan deskripsi mismatch dan jumlah mismatch per cabang/wilayah
        mismatch_summary = (
            df_selected_IC.groupby("Cabang")
            .agg(
                Jumlah_Mismatch_Limit=("Mismatch_Desc", lambda x: x.notna().sum()),
                Mismatch_Limit_Detail=(
                    "Mismatch_Desc",
                    lambda x: "; ".join([d for d in x.dropna()]),
                ),
            )
            .reset_index()
        )

        # Merge ke summary sekali saja
        df_summary = df_summary.merge(
            mismatch_summary, left_on="Wilayah", right_on="Cabang", how="left"
        )

        # --- 8. Bersihkan kolom duplikat cabang ---
        df_summary.drop(
            columns=["Cabang", "CABANG", "Cabang_x", "Cabang_y"],
            inplace=True,
            errors="ignore",
        )

        # --- 9. Tampilkan Summary Gabungan ---
        st.subheader("Summary Balance Checking ABACUS")
        st.dataframe(
            df_summary.style.format(
                {
                    "Awal_saldo": "{:,.0f}",
                    "Supply": "{:,.0f}",
                    "Total Saldo": "{:,.0f}",
                    "Total nominal rencana pengisian": "{:,.0f}",
                    "Nominal realisasi pengisian": "{:,.0f}",
                    "Nominal sisa uang": "{:,.0f}",
                    "Saldo Akhir Rencana": "{:,.0f}",
                    "Saldo Akhir Realisasi": "{:,.0f}",
                    "Posisi Saldo": "{:,.0f}",
                    "Selisih akhir Rencana": "{:,.0f}",
                    "Selisih akhir Realisasi": "{:,.0f}",
                    "Selisih Jumlah Replenishment": "{:,.0f}",
                }
            )
        )
