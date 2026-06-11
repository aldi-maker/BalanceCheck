import streamlit as st
import pandas as pd
from openpyxl import load_workbook


def run():
    st.title("Laporan RPL & ATM ADVANTAGE")

    # --- Upload terpisah ---
    file_CIT = st.file_uploader("Upload laporan CIT (sheet CIT)", type=["xlsx"])
    file_IC = st.file_uploader("Upload laporan IC (sheet IC)", type=["xlsx"])
    file_wilayah = st.file_uploader(
        "Upload laporan Rencana & Wilayah(saldo Posisi Saldo)", type=["xlsx"]
    )
    file_PosisiAwal = st.file_uploader("Upload laporan Saldo awal", type=["xlsx"])

    if file_CIT and file_IC and file_wilayah and file_PosisiAwal:
        # --- 0. Detail CIT ---
        df_CIT = pd.read_excel(file_CIT)

        cols_map = {
            "WSID_FINAL": "ATM ID",
            "ATMName": "Nama ATM",
            "DenomValue": "Denom",
            "DateReplenish": "Tanggal Replenish",
            "FinishedTimeATM": "JamRPL",
            "BranchName": "Cabang",
            "CassetteQty_Previous": "Total Kaset Sebelumnya",
            "QtyPerCassette_Previous": "Admin Sebelumnya",
            "Cassette1RemainingQty": "Sisa uang Kaset 1",
            "Cassette2RemainingQty": "Sisa uang Kaset 2",
            "Cassette3RemainingQty": "Sisa uang Kaset 3",
            "Cassette4RemainingQty": "Sisa uang Kaset 4",
            "RejectQty": "Sisa uang Kaset Reject",
            "TotalDispanse": "Total uang keluar",
            "CassetteQty_Current": "Total Kaset Terbaru",
            "QtyPerCassette_Current": "Admin Terbaru",
        }

        available_cols = [c for c in cols_map.keys() if c in df_CIT.columns]
        df_selected_CIT = df_CIT[available_cols].copy()
        df_selected_CIT.rename(columns=cols_map, inplace=True)

        # Konversi numerik
        for col in df_selected_CIT.columns:
            df_selected_CIT[col] = pd.to_numeric(df_selected_CIT[col], errors="ignore")

        # Hitung turunan
        df_selected_CIT["Total Sisa Uang"] = (
            df_selected_CIT.get("Sisa uang Kaset 1", 0)
            + df_selected_CIT.get("Sisa uang Kaset 2", 0)
            + df_selected_CIT.get("Sisa uang Kaset 3", 0)
            + df_selected_CIT.get("Sisa uang Kaset 4", 0)
            + df_selected_CIT.get("Sisa uang Kaset Reject", 0)
        ) * df_selected_CIT.get("Denom", 1)
        df_selected_CIT["Total Amount Uang Keluar"] = df_selected_CIT.get(
            "Total uang keluar", 0
        ) * df_selected_CIT.get("Denom", 1)

        # st.subheader("Detail ATM (CIT)")
        # st.dataframe(df_selected_CIT.head(10))

        # --- 1. Detail IC ---
        xls_IC = pd.ExcelFile(file_IC)
        sheet_names = xls_IC.sheet_names

        # Tampilkan pilihan ke user
        selected_sheet = st.selectbox(
            "Pilih tanggal dari laporan IC yang mau dibaca:", sheet_names
        )
        df_IC = pd.read_excel(file_IC, sheet_name=selected_sheet, skiprows=1)
        df_IC.rename(
            columns={
                "Unnamed: 0": "NO",
                "Unnamed: 1": "ID ATM",
                "Unnamed: 2": "Nama Lokasi",
            },
            inplace=True,
        )

        selected_cols = [
            "ID ATM",
            "Nama Lokasi",
            "Cabang",
            "Denom",
            "Limit",
            "Saldo",
            "Nominal",
        ]
        df_selected_IC = df_IC[[c for c in selected_cols if c in df_IC.columns]]

        for col in ["Denom", "Limit", "Saldo", "Nominal"]:
            if col in df_selected_IC.columns:
                df_selected_IC[col] = pd.to_numeric(
                    df_selected_IC[col], errors="coerce"
                )

        # st.subheader("Preview Data ATM (IC)")
        # st.dataframe(df_selected_IC.head(10))

        # --- 2. Rekap per Cabang (Rencana) ---
        df_cabang = pd.read_excel(file_wilayah, sheet_name="Sheet1", skiprows=1)
        cols_order = ["WSID", "LOKASI", "CABANG", "DENOM", "LIMIT", "KETERANGAN"]
        df_selected_cabang = df_cabang[
            [c for c in cols_order if c in df_cabang.columns]
        ]

        summary_cabang = (
            df_selected_cabang.groupby("CABANG")
            .agg(Jumlah_RPL=("WSID", "count"), Total_Limit=("LIMIT", "sum"))
            .reset_index()
        )

        # --- 3. Rekap per Wilayah ---
        wb = load_workbook(file_wilayah, data_only=True)
        ws = wb.active
        mapping = {
            "KARAWANG": ["D4", "E4", "F4", "D5", "E5", "F5"],
            "MERUYA": ["G4", "H4", "I4", "G5", "H5", "I5"],
            "RAWAMANGUN": ["J4", "K4", "L4", "J5", "K5", "L5"],
            "SERANG": ["M4", "N4", "O4", "M5", "N5", "O5"],
            "BANDUNG": ["P4", "Q4", "R4", "P5", "Q5", "R5"],
            "CIREBON": ["D12", "E12", "F12", "D13", "E13", "F13"],
            "TASIKMALAYA": ["G12", "H12", "I12", "G13", "H13", "I13"],
            "PURWOKERTO": ["J12", "K12", "L12", "J13", "K13", "L13"],
            "SOLO": ["M12", "N12", "O12", "M13", "N13", "O13"],
            "YOGYAKARTA": ["P12", "Q12", "R12", "P13", "Q13", "R13"],
            "SEMARANG": ["D20", "E20", "F20", "D21", "E21", "F21"],
            "TEGAL": ["G20", "H20", "I20", "G21", "H21", "I21"],
            "KUDUS": ["J20", "K20", "L20", "J21", "K21", "L21"],
            "JEMBER": ["M20", "N20", "O20", "M21", "N21", "O21"],
            "KEDIRI": ["P20", "Q20", "R20", "P21", "Q21", "R21"],
            "MALANG": ["D28", "E28", "F28", "D29", "E29", "F29"],
            "SURABAYA": ["G28", "H28", "I28", "G29", "H29", "I29"],
            "LAMPUNG": ["J28", "K28", "L28", "J29", "K29", "L29"],
            "BENGKULU": ["M28", "N28", "O28", "M29", "N29", "O29"],
            "JAMBI": ["P28", "Q28", "R28", "P29", "Q29", "R29"],
            "PALEMBANG": ["D36", "E36", "F36", "D37", "E37", "F37"],
            "BATAM": ["G36", "H36", "I36", "G37", "H37", "I37"],
            "MEDAN": ["J36", "K36", "L36", "J37", "K37", "L37"],
            "PEKANBARU": ["M36", "N36", "O36", "M37", "N37", "O37"],
            "PONTIANAK": ["P36", "Q36", "R36", "P37", "Q37", "R37"],
            "BALIKPAPAN": ["D44", "E44", "F44", "D45", "E45", "F45"],
            "BANJARMASIN": ["G44", "H44", "I44", "G45", "H45", "I45"],
            "SAMARINDA": ["J44", "K44", "L44", "J45", "K45", "L45"],
            "SINGKAWANG": ["M44", "N44", "O44", "M45", "N45", "O45"],
            "MANADO": ["P44", "Q44", "R44", "P45", "Q45", "R45"],
            "MATARAM": ["D52", "E52", "F52", "D53", "E53", "F53"],
            "KUPANG": ["G52", "H52", "I52", "G53", "H53", "I53"],
            "DENPASAR": ["J52", "K52", "L52", "J53", "K53", "L53"],
        }
        data = []
        for wilayah, cells in mapping.items():
            values = [ws[c].value or 0 for c in cells]
            total_d100 = values[0] + values[3]
            total_d50 = values[1] + values[4]
            total_d20 = values[2] + values[5]
            awal_saldo = values[0] + values[1] + values[2]
            supply = values[3] + values[4] + values[5]
            total_saldo = total_d100 + total_d50 + total_d20
            row = {
                "Wilayah": wilayah,
                "Awal_saldo": awal_saldo,
                "Supply": supply,
                "Total Saldo": total_saldo,
            }
            data.append(row)
        df_wilayah = pd.DataFrame(data)

        # --- 3.1 Rekap posisi awal ---
        wb = load_workbook(file_PosisiAwal, data_only=True)
        ws = wb.active
        mapping = {
            "KARAWANG": ["D4", "E4", "F4"],
            "MERUYA": ["G4", "H4", "I4"],
            "RAWAMANGUN": ["J4", "K4", "L4"],
            "SERANG": ["M4", "N4", "O4"],
            "BANDUNG": ["P4", "Q4", "R4"],
            "CIREBON": ["D12", "E12", "F12"],
            "TASIKMALAYA": ["G12", "H12", "I12"],
            "PURWOKERTO": ["J12", "K12", "L12"],
            "SOLO": ["M12", "N12", "O12"],
            "YOGYAKARTA": ["P12", "Q12", "R12"],
            "SEMARANG": ["D20", "E20", "F20"],
            "TEGAL": ["G20", "H20", "I20"],
            "KUDUS": ["J20", "K20", "L20"],
            "JEMBER": ["M20", "N20", "O20"],
            "KEDIRI": ["P20", "Q20", "R20"],
            "MALANG": ["D28", "E28", "F28"],
            "SURABAYA": ["G28", "H28", "I28"],
            "LAMPUNG": ["J28", "K28", "L28"],
            "BENGKULU": ["M28", "N28", "O28"],
            "JAMBI": ["P28", "Q28", "R28"],
            "PALEMBANG": ["D36", "E36", "F36"],
            "BATAM": ["G36", "H36", "I36"],
            "MEDAN": ["J36", "K36", "L36"],
            "PEKANBARU": ["M36", "N36", "O36"],
            "PONTIANAK": ["P36", "Q36", "R36"],
            "BALIKPAPAN": ["D44", "E44", "F44"],
            "BANJARMASIN": ["G44", "H44", "I44"],
            "SAMARINDA": ["J44", "K44", "L44"],
            "SINGKAWANG": ["M44", "N44", "O44"],
            "MANADO": ["P44", "Q44", "R44"],
            "MATARAM": ["D52", "E52", "F52"],
            "KUPANG": ["G52", "H52", "I52"],
            "DENPASAR": ["J52", "K52", "L52"],
        }
        data_realisasi = []
        for wilayah, cells in mapping.items():
            values = [ws[c].value or 0 for c in cells]
            total_d100 = values[0]
            total_d50 = values[1]
            total_d20 = values[2]
            posisi_saldo = values[0] + values[1] + values[2]
            row = {"Wilayah": wilayah, "Posisi Saldo": posisi_saldo}
            data_realisasi.append(row)
        df_realisasi = pd.DataFrame(data_realisasi)

        # --- 4. Gabungkan Wilayah dengan Rekap Cabang ---
        df_summary = df_wilayah.merge(
            summary_cabang, left_on="Wilayah", right_on="CABANG", how="left"
        )
        df_summary.rename(
            columns={
                "Jumlah_RPL": "Jumlah lokasi rencana pengisian",
                "Total_Limit": "Total nominal rencana pengisian",
            },
            inplace=True,
        )

        # --- 5. Tambahkan kolom realisasi dari IC ---
        summary_realisasi = (
            df_selected_IC.groupby("Cabang")
            .agg(
                Jumlah_Realisasi=("Nominal", lambda x: (x > 0).sum()),
                Nominal_Realisasi=("Nominal", "sum"),
            )
            .reset_index()
        )
        df_summary = df_summary.merge(
            summary_realisasi, left_on="Wilayah", right_on="Cabang", how="left"
        )
        df_summary.rename(
            columns={
                "Jumlah_Realisasi": "Jumlah lokasi realisasi pengisian",
                "Nominal_Realisasi": "Nominal realisasi pengisian",
            },
            inplace=True,
        )

        # --- 6. Tambahkan kolom Sisa uang dari CIT ---
        summary_sisa = (
            df_selected_CIT.groupby("Cabang")
            .agg(
                Jumlah_SisaUang=("Total Sisa Uang", lambda x: (x > 0).sum()),
                Nominal_SisaUang=("Total Sisa Uang", "sum"),
            )
            .reset_index()
        )
        df_summary = df_summary.merge(
            summary_sisa, left_on="Wilayah", right_on="Cabang", how="left"
        )
        df_summary.rename(
            columns={
                "Jumlah_SisaUang": "Jumlah lokasi sisa uang",
                "Nominal_SisaUang": "Nominal sisa uang",
            },
            inplace=True,
        )

        # --- 7. Hitung Saldo Akhir Rencana ---
        df_summary["Saldo Akhir Rencana"] = (
            df_summary["Total Saldo"].fillna(0)
            - df_summary["Total nominal rencana pengisian"].fillna(0)
            + df_summary["Nominal sisa uang"].fillna(0)
        )

        # --- 7.1 Hitung Saldo Akhir Realisasi ---
        df_summary["Saldo Akhir Realisasi"] = (
            df_summary["Total Saldo"].fillna(0)
            - df_summary["Nominal realisasi pengisian"].fillna(0)
            + df_summary["Nominal sisa uang"].fillna(0)
        )

        # --- 7.2 Hitung Selisih nominal Rencana RPL ---
        df_summary = df_summary.merge(df_realisasi, on="Wilayah", how="left")
        df_summary["Selisih akhir Rencana"] = df_summary["Posisi Saldo"].fillna(
            0
        ) - df_summary["Saldo Akhir Rencana"].fillna(0)

        # --- 7.3 Hitung Selisih nominal realisasi RPL ---
        df_summary["Selisih akhir Realisasi"] = df_summary["Posisi Saldo"].fillna(
            0
        ) - df_summary["Saldo Akhir Realisasi"].fillna(0)

        # --- 7.4 Hitung Selisih Jumlah Replenishment ---
        df_summary["Selisih Jumlah Replenishment"] = df_summary[
            "Jumlah lokasi realisasi pengisian"
        ].fillna(0) - df_summary["Jumlah lokasi rencana pengisian"].fillna(0)

        # --- 7.5 Tambahkan kolom Remark per Wilayah ---
        # Pastikan WSID dan ID ATM jadi string tanpa .0
        wsid_per_cabang = (
            df_selected_cabang.groupby("CABANG")["WSID"]
            .apply(lambda x: set(x.dropna().astype(int).astype(str)))
            .to_dict()
        )
        idatm_per_cabang = (
            df_selected_IC.groupby("Cabang")["ID ATM"]
            .apply(lambda x: set(x.dropna().astype(int).astype(str)))
            .to_dict()
        )

        # Pastikan ID ATM di df_selected_IC jadi string bersih
        df_selected_IC["ID ATM"] = (
            df_selected_IC["ID ATM"]
            .astype(str)
            .str.strip()
            .str.replace(r"\.0$", "", regex=True)
        )

        remarks = []
        for idx, row in df_summary.iterrows():
            wilayah = row["Wilayah"]
            wsid_set = wsid_per_cabang.get(wilayah, set())
            idatm_set = idatm_per_cabang.get(wilayah, set())

            wsid_not_in_ic = wsid_set - idatm_set
            idatm_not_in_rencana = idatm_set - wsid_set

            if wsid_not_in_ic or idatm_not_in_rencana:
                remark_text = []
                if wsid_not_in_ic:
                    remark_text.append(
                        "WSID not in IC: " + ", ".join(sorted(wsid_not_in_ic))
                    )
                if idatm_not_in_rencana:
                    details = []
                    for atm_id in sorted(idatm_not_in_rencana):
                        row_ic = df_selected_IC.loc[df_selected_IC["ID ATM"] == atm_id]
                        if not row_ic.empty:
                            nominal_val = (
                                row_ic["Nominal"].iloc[0]
                                if pd.notna(row_ic["Nominal"].iloc[0])
                                else 0
                            )
                            details.append(f"{atm_id} (Nominal {int(nominal_val):,})")
                        else:
                            details.append(atm_id)
                    remark_text.append("ID ATM not in Rencana: " + ", ".join(details))
                remarks.append("; ".join(remark_text))
            else:
                remarks.append("Match")

        df_summary["Remark"] = remarks

        # --- 7.6 Hitung mismatch limit per terminal dari Detail IC ---
        df_selected_IC["Mismatch_Limit"] = df_selected_IC["Nominal"].fillna(
            0
        ) - df_selected_IC["Limit"].fillna(0)

        # Buat deskripsi mismatch per ATM (hindari NaN dengan pengecekan)
        def make_desc(row):
            if pd.notna(row["ID ATM"]) and row["Mismatch_Limit"] != 0:
                try:
                    atm_id = str(int(float(row["ID ATM"])))  # aman dari float NaN
                except Exception:
                    atm_id = str(row["ID ATM"])
                nominal_val = int(row["Nominal"]) if pd.notna(row["Nominal"]) else 0
                limit_val = int(row["Limit"]) if pd.notna(row["Limit"]) else 0
                return f"ATM {atm_id} mismatch: Realisasi {nominal_val:,} vs Limit {limit_val:,}"
            return None

        df_selected_IC["Mismatch_Desc"] = df_selected_IC.apply(make_desc, axis=1)

        # Gabungkan deskripsi mismatch dan jumlah mismatch per cabang/wilayah
        mismatch_summary = (
            df_selected_IC.groupby("Cabang")
            .agg(
                Jumlah_Mismatch_Limit=("Mismatch_Desc", lambda x: x.notna().sum()),
                Mismatch_Limit_Detail=(
                    "Mismatch_Desc",
                    lambda x: "; ".join([d for d in x.dropna()]),
                ),
            )
            .reset_index()
        )

        # Merge ke summary sekali saja
        df_summary = df_summary.merge(
            mismatch_summary, left_on="Wilayah", right_on="Cabang", how="left"
        )

        # --- 8. Bersihkan kolom duplikat cabang ---
        df_summary.drop(
            columns=["Cabang", "CABANG", "Cabang_x", "Cabang_y"],
            inplace=True,
            errors="ignore",
        )

        # --- 8.1 Pastikan kolom jumlah lokasi jadi integer ---
        for col in [
            "Jumlah lokasi rencana pengisian",
            "Jumlah lokasi realisasi pengisian",
            "Jumlah lokasi sisa uang",
            "Jumlah_Mismatch_Limit",
        ]:
            if col in df_summary.columns:
                df_summary[col] = (
                    pd.to_numeric(df_summary[col], errors="coerce")
                    .fillna(0)
                    .astype(int)
                )

        # --- 9. Tampilkan Summary Gabungan ---
        st.subheader("Summary Balance Checking ADVANTAGE")
        st.dataframe(
            df_summary.style.format(
                {
                    "Awal_saldo": "{:,.0f}",
                    "Supply": "{:,.0f}",
                    "Total Saldo": "{:,.0f}",
                    "Total nominal rencana pengisian": "{:,.0f}",
                    "Nominal realisasi pengisian": "{:,.0f}",
                    "Nominal sisa uang": "{:,.0f}",
                    "Saldo Akhir Rencana": "{:,.0f}",
                    "Saldo Akhir Realisasi": "{:,.0f}",
                    "Posisi Saldo": "{:,.0f}",
                    "Selisih akhir Rencana": "{:,.0f}",
                    "Selisih akhir Realisasi": "{:,.0f}",
                    "Selisih Jumlah Replenishment": "{:,.0f}",
                }
            )
        )
